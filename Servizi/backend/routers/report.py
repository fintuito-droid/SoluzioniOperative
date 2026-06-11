"""
routers/report.py — Report designer ed export
=============================================
- CRUD modelli report (JSON in report_templates) — admin e responsabile
- Render PDF dal modello (reportlab): bande intestazione/piè con elementi
  liberi posizionati in mm, tabella dati centrale paginata
- Export Excel a layout fisso (openpyxl) per le 3 sorgenti

Sorgenti dati: presenze, monte_ore, riepilogo (per postazione).
"""

import io
import json
from datetime import datetime, date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Body
from fastapi.responses import StreamingResponse

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.pdfgen.canvas import Canvas

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from db.database import db
from auth import get_current_user, require_role
from routers.presenze import lista_presenze, monte_ore

# Tutta la sezione Report è riservata agli amministratori
router = APIRouter(prefix="/report", tags=["Report"],
                   dependencies=[Depends(require_role("admin"))])

RUOLI_DESIGNER = ("admin",)


# ═════════════════════════════ Catalogo sorgenti ═════════════════════════════

CAMPI_SORGENTE = {
    "presenze": [
        {"campo": "data_servizio", "etichetta": "Data",        "formato": "data"},
        {"campo": "nominativo",    "etichetta": "Dipendente"},
        {"campo": "qualifica",     "etichetta": "Qualifica"},
        {"campo": "funzione",      "etichetta": "Funzione"},
        {"campo": "postazione",    "etichetta": "Postazione"},
        {"campo": "orario_inizio", "etichetta": "Inizio"},
        {"campo": "orario_fine",   "etichetta": "Fine"},
        {"campo": "ore_totali",    "etichetta": "Ore"},
        {"campo": "stato",         "etichetta": "Stato"},
        {"campo": "note_consuntivo", "etichetta": "Note"},
    ],
    "monte_ore": [
        {"campo": "cognome",      "etichetta": "Cognome"},
        {"campo": "nome",         "etichetta": "Nome"},
        {"campo": "qualifica",    "etichetta": "Qualifica"},
        {"campo": "turni_totali", "etichetta": "Turni"},
        {"campo": "ore_totali",   "etichetta": "Ore totali"},
    ],
    "riepilogo": [
        {"campo": "postazione",   "etichetta": "Postazione"},
        {"campo": "turni_totali", "etichetta": "Turni"},
        {"campo": "ore_totali",   "etichetta": "Ore totali"},
        {"campo": "dipendenti",   "etichetta": "Dipendenti"},
    ],
}

# Campi dinamici utilizzabili negli elementi 'campo' delle bande
CAMPI_INTESTAZIONE = ["sottotitolo", "campagna", "postazione", "periodo", "utente"]


@router.get("/sorgenti")
def get_sorgenti(current_user: dict = Depends(get_current_user)):
    """Catalogo campi per il designer."""
    return {"sorgenti": CAMPI_SORGENTE, "campi_intestazione": CAMPI_INTESTAZIONE}


# ═════════════════════════════ Dati ══════════════════════════════════════════

def _fmt_data_it(v) -> str:
    s = str(v)[:10]
    parti = s.split("-")
    return f"{parti[2]}/{parti[1]}/{parti[0]}" if len(parti) == 3 else s


def _carica_dati(sorgente: str, filtri: dict, current_user: dict) -> list[dict]:
    if sorgente == "presenze":
        rows = lista_presenze(
            campagna_id=filtri.get("campagna_id"),
            personale_id=None,
            postazione_id=filtri.get("postazione_id"),
            data_da=filtri.get("data_da"),
            data_a=filtri.get("data_a"),
            stato=filtri.get("stato"),
            current_user=current_user,
        )
        for r in rows:
            r["nominativo"] = f"{r.get('cognome') or ''} {r.get('nome_dip') or ''}".strip()
        return rows

    if sorgente == "monte_ore":
        return monte_ore(
            campagna_id=filtri.get("campagna_id"),
            data_da=filtri.get("data_da"),
            data_a=filtri.get("data_a"),
            postazione_id=filtri.get("postazione_id"),
            current_user=current_user,
        )

    if sorgente == "riepilogo":
        rows = lista_presenze(
            campagna_id=filtri.get("campagna_id"),
            personale_id=None,
            postazione_id=filtri.get("postazione_id"),
            data_da=filtri.get("data_da"),
            data_a=filtri.get("data_a"),
            stato=filtri.get("stato"),
            current_user=current_user,
        )
        agg: dict[str, dict] = {}
        for r in rows:
            key = r.get("postazione") or "—"
            a = agg.setdefault(key, {"postazione": key, "turni_totali": 0,
                                     "ore_totali": 0.0, "_pers": set()})
            a["turni_totali"] += 1
            a["ore_totali"]   += r.get("ore_totali") or 0
            a["_pers"].add(r.get("personale_id"))
        out = []
        for a in sorted(agg.values(), key=lambda x: x["postazione"]):
            a["dipendenti"] = len(a.pop("_pers"))
            a["ore_totali"] = round(a["ore_totali"], 1)
            out.append(a)
        return out

    raise HTTPException(422, f"Sorgente '{sorgente}' sconosciuta")


def _contesto(filtri: dict, current_user: dict) -> dict:
    """Valori per i campi dinamici delle bande."""
    campagna = postazione = ""
    if filtri.get("campagna_id"):
        c = db.fetch_one("SELECT anno FROM campagne_aib WHERE id=?", (filtri["campagna_id"],))
        if c:
            campagna = f"Campagna AIB {c['anno']}"
    if filtri.get("postazione_id"):
        p = db.fetch_one("SELECT codice FROM postazioni WHERE id=?", (filtri["postazione_id"],))
        if p:
            postazione = f"Postazione {p['codice']}"
    periodo = ""
    if filtri.get("data_da") or filtri.get("data_a"):
        da = _fmt_data_it(filtri["data_da"]) if filtri.get("data_da") else "…"
        a  = _fmt_data_it(filtri["data_a"])  if filtri.get("data_a")  else "…"
        periodo = f"{da} – {a}"
    sottotitolo = " · ".join(x for x in (postazione, periodo, campagna) if x)
    return {
        "sottotitolo": sottotitolo,
        "campagna":    campagna,
        "postazione":  postazione,
        "periodo":     periodo,
        "utente":      current_user.get("username", ""),
    }


def _filtri_da_query(campagna_id, postazione_id, data_da, data_a, stato) -> dict:
    return {"campagna_id": campagna_id, "postazione_id": postazione_id,
            "data_da": data_da, "data_a": data_a, "stato": stato}


# ═════════════════════════════ Render PDF ════════════════════════════════════

def _colore(hex_str, default="#000000"):
    try:
        return HexColor(hex_str or default)
    except Exception:
        return HexColor(default)


def _disegna_elemento(c: Canvas, el: dict, ox: float, oy: float, banda_h: float,
                      contesto: dict, pagina: int):
    """
    Disegna un elemento della banda. (ox, oy) = origine banda in punti PDF
    (oy = bordo SUPERIORE della banda). Le coordinate elemento sono in mm
    dal vertice alto-sinistro della banda.
    """
    tipo = el.get("tipo", "testo")
    x = ox + float(el.get("x", 0)) * mm
    y = oy - float(el.get("y", 0)) * mm          # y cresce verso il basso
    w = float(el.get("w", 50)) * mm
    f = el.get("font") or {}
    size   = float(f.get("size", 9))
    bold   = bool(f.get("bold"))
    italic = bool(f.get("italic"))
    colore = _colore(f.get("colore"))
    align  = el.get("align", "left")

    if tipo == "linea":
        c.setStrokeColor(colore)
        c.setLineWidth(max(float(el.get("h", 0)) or 0.7, 0.5))
        c.line(x, y, x + w, y)
        return

    if tipo == "rettangolo":
        h = float(el.get("h", 10)) * mm
        c.setStrokeColor(colore)
        c.setLineWidth(0.7)
        c.rect(x, y - h, w, h, stroke=1, fill=0)
        return

    # Elementi testuali
    if tipo == "testo":
        testo = el.get("testo", "")
    elif tipo == "campo":
        testo = contesto.get(el.get("campo", ""), "")
    elif tipo == "data":
        testo = datetime.now().strftime("%d/%m/%Y")
    elif tipo == "numpagina":
        testo = f"Pag. {pagina}"
    else:
        return

    nome_font = "Helvetica"
    if bold and italic: nome_font = "Helvetica-BoldOblique"
    elif bold:          nome_font = "Helvetica-Bold"
    elif italic:        nome_font = "Helvetica-Oblique"

    c.setFont(nome_font, size)
    c.setFillColor(colore)
    ty = y - size  # baseline sotto il bordo alto dell'elemento
    if align == "center":
        c.drawCentredString(x + w / 2, ty, str(testo))
    elif align == "right":
        c.drawRightString(x + w, ty, str(testo))
    else:
        c.drawString(x, ty, str(testo))


def render_pdf(definizione: dict, rows: list[dict], contesto: dict) -> bytes:
    pagina_cfg = definizione.get("pagina", {})
    margini    = pagina_cfg.get("margini", {})
    m_sx, m_dx  = float(margini.get("sx", 15)) * mm, float(margini.get("dx", 15)) * mm
    m_up, m_dn  = float(margini.get("alto", 12)) * mm, float(margini.get("basso", 12)) * mm

    page = A4 if pagina_cfg.get("orientamento", "portrait") == "portrait" else landscape(A4)
    PW, PH = page

    intest = definizione.get("intestazione", {"altezza": 0, "elementi": []})
    pie    = definizione.get("pie",          {"altezza": 0, "elementi": []})
    h_int  = float(intest.get("altezza", 0)) * mm
    h_pie  = float(pie.get("altezza", 0)) * mm

    tab     = definizione.get("tabella", {})
    colonne = tab.get("colonne", [])
    stile   = tab.get("stile", {})
    fs      = float(stile.get("font_size", 8))
    riga_h  = fs * 1.9
    col_hdr = _colore(stile.get("colore_header"), "#C0392B")
    col_txt = _colore(stile.get("testo_header"), "#FFFFFF")
    zebra   = stile.get("zebra", True)
    bordi   = stile.get("bordi", True)

    larg_tot = sum(float(col.get("larghezza", 20)) for col in colonne) or 1
    area_w   = PW - m_sx - m_dx
    scala    = area_w / (larg_tot * mm)          # adatta le colonne all'area utile
    widths   = [float(col.get("larghezza", 20)) * mm * scala for col in colonne]

    buf = io.BytesIO()
    c = Canvas(buf, pagesize=page)
    pagina = 1

    def bande():
        _y_int = PH - m_up
        for el in intest.get("elementi", []):
            _disegna_elemento(c, el, m_sx, _y_int, h_int, contesto, pagina)
        _y_pie = m_dn + h_pie
        for el in pie.get("elementi", []):
            _disegna_elemento(c, el, m_sx, _y_pie, h_pie, contesto, pagina)

    def header_tabella(y):
        c.setFillColor(col_hdr)
        c.rect(m_sx, y - riga_h, area_w, riga_h, stroke=0, fill=1)
        c.setFillColor(col_txt)
        c.setFont("Helvetica-Bold", fs)
        x = m_sx
        for col, w in zip(colonne, widths):
            c.drawString(x + 2, y - riga_h + fs * 0.45, str(col.get("etichetta", col.get("campo", ""))))
            x += w
        return y - riga_h

    def valore(col, r):
        v = r.get(col.get("campo", ""), "")
        if v is None:
            return ""
        if col.get("formato") == "data":
            return _fmt_data_it(v)
        if isinstance(v, float):
            return f"{v:g}"
        return str(v)

    y_inizio = PH - m_up - h_int - 4
    y_fine   = m_dn + h_pie + 4

    bande()
    y = header_tabella(y_inizio)
    idx = 0
    for r in rows:
        if y - riga_h < y_fine:
            c.showPage()
            pagina += 1
            bande()
            y = header_tabella(y_inizio)
        if zebra and idx % 2 == 1:
            c.setFillColor(HexColor("#F5F5F5"))
            c.rect(m_sx, y - riga_h, area_w, riga_h, stroke=0, fill=1)
        if bordi:
            c.setStrokeColor(HexColor("#DDDDDD"))
            c.setLineWidth(0.4)
            c.line(m_sx, y - riga_h, m_sx + area_w, y - riga_h)
        c.setFillColor(HexColor("#222222"))
        c.setFont("Helvetica", fs)
        x = m_sx
        for col, w in zip(colonne, widths):
            testo = valore(col, r)
            al = col.get("align", "left")
            ty = y - riga_h + fs * 0.45
            if al == "right":
                c.drawRightString(x + w - 2, ty, testo)
            elif al == "center":
                c.drawCentredString(x + w / 2, ty, testo)
            else:
                c.drawString(x + 2, ty, testo)
            x += w
        y -= riga_h
        idx += 1

    if not rows:
        c.setFont("Helvetica-Oblique", 10)
        c.setFillColor(HexColor("#888888"))
        c.drawString(m_sx, y - 14, "Nessun dato per i filtri selezionati.")

    c.showPage()
    c.save()
    return buf.getvalue()


# ═════════════════════════════ Excel fisso ═══════════════════════════════════

def _excel(titolo: str, intestazioni: list[str], righe: list[list], widths: list[int]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = titolo[:31]

    rosso  = "C0392B"
    bordo  = Border(bottom=Side(style="thin", color="DDDDDD"))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(intestazioni))
    cella = ws.cell(row=1, column=1, value=titolo)
    cella.font = Font(size=14, bold=True, color=rosso)

    for j, h in enumerate(intestazioni, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=rosso)
        c.alignment = Alignment(horizontal="center")

    for i, riga in enumerate(righe, start=4):
        for j, v in enumerate(riga, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = bordo

    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=3, column=j).column_letter].width = w

    if righe:
        ws.auto_filter.ref = f"A3:{ws.cell(row=3, column=len(intestazioni)).coordinate}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_per_sorgente(sorgente: str, rows: list[dict], contesto: dict) -> tuple[bytes, str]:
    sub = contesto.get("sottotitolo", "")
    if sorgente == "presenze":
        titolo = f"Registro presenze {sub}".strip()
        dati = [[_fmt_data_it(r.get("data_servizio")), r.get("nominativo"), r.get("qualifica"),
                 r.get("funzione"), r.get("postazione"), r.get("orario_inizio"),
                 r.get("orario_fine"), r.get("ore_totali"), r.get("stato")] for r in rows]
        return _excel(titolo,
                      ["Data", "Dipendente", "Qualifica", "Funzione", "Postazione",
                       "Inizio", "Fine", "Ore", "Stato"],
                      dati, [12, 28, 10, 16, 14, 8, 8, 8, 14]), "presenze.xlsx"
    if sorgente == "monte_ore":
        titolo = f"Monte ore {sub}".strip()
        dati = [[r.get("cognome"), r.get("nome"), r.get("qualifica"),
                 r.get("turni_totali"), r.get("ore_totali")] for r in rows]
        return _excel(titolo, ["Cognome", "Nome", "Qualifica", "Turni", "Ore totali"],
                      dati, [24, 24, 14, 10, 12]), "monte_ore.xlsx"
    if sorgente == "riepilogo":
        titolo = f"Riepilogo campagna {sub}".strip()
        dati = [[r.get("postazione"), r.get("turni_totali"), r.get("ore_totali"),
                 r.get("dipendenti")] for r in rows]
        return _excel(titolo, ["Postazione", "Turni", "Ore totali", "Dipendenti"],
                      dati, [24, 12, 12, 12]), "riepilogo.xlsx"
    raise HTTPException(422, f"Sorgente '{sorgente}' sconosciuta")


# ═════════════════════════════ Endpoint modelli ══════════════════════════════

def _template_out(row: dict) -> dict:
    out = dict(row)
    try:
        out["definizione"] = json.loads(row.get("definizione") or "{}")
    except Exception:
        out["definizione"] = {}
    return out


@router.get("/templates")
def lista_templates(current_user: dict = Depends(get_current_user)):
    rows = db.fetch_all("SELECT * FROM report_templates ORDER BY nome")
    return [_template_out(r) for r in rows]


@router.get("/templates/{tid}")
def get_template(tid: int, current_user: dict = Depends(get_current_user)):
    row = db.fetch_one("SELECT * FROM report_templates WHERE id=?", (tid,))
    if not row:
        raise HTTPException(404, "Modello non trovato")
    return _template_out(row)


@router.post("/templates", dependencies=[Depends(require_role(*RUOLI_DESIGNER))])
def crea_template(body: dict, current_user: dict = Depends(get_current_user)):
    if not body.get("nome") or not body.get("sorgente"):
        raise HTTPException(422, "Nome e sorgente sono obbligatori")
    new_id = db.execute(
        "INSERT INTO report_templates ([nome],[sorgente],[definizione],[creato_da],[creato_il]) VALUES (?,?,?,?,?)",
        (body["nome"], body["sorgente"],
         json.dumps(body.get("definizione") or {}, ensure_ascii=False),
         current_user["id"], datetime.now())
    )
    return get_template(new_id, current_user)


@router.put("/templates/{tid}", dependencies=[Depends(require_role(*RUOLI_DESIGNER))])
def aggiorna_template(tid: int, body: dict, current_user: dict = Depends(get_current_user)):
    row = db.fetch_one("SELECT id FROM report_templates WHERE id=?", (tid,))
    if not row:
        raise HTTPException(404, "Modello non trovato")
    db.execute(
        "UPDATE report_templates SET [nome]=?, [definizione]=?, [aggiornato_il]=? WHERE id=?",
        (body.get("nome", ""),
         json.dumps(body.get("definizione") or {}, ensure_ascii=False),
         datetime.now(), tid)
    )
    return get_template(tid, current_user)


@router.delete("/templates/{tid}", dependencies=[Depends(require_role(*RUOLI_DESIGNER))])
def elimina_template(tid: int):
    n = db.execute("DELETE FROM report_templates WHERE id=?", (tid,))
    if n == 0:
        raise HTTPException(404, "Modello non trovato")
    return {"ok": True}


# ═════════════════════════════ Endpoint generazione ══════════════════════════

@router.get("/pdf/{tid}")
def genera_pdf(
    tid: int,
    campagna_id:   Optional[int]  = Query(None),
    postazione_id: Optional[int]  = Query(None),
    data_da:       Optional[date] = Query(None),
    data_a:        Optional[date] = Query(None),
    stato:         Optional[str]  = Query(None),
    current_user:  dict           = Depends(get_current_user),
):
    tpl = get_template(tid, current_user)
    filtri = _filtri_da_query(campagna_id, postazione_id, data_da, data_a, stato)
    rows = _carica_dati(tpl["sorgente"], filtri, current_user)
    ctx  = _contesto(filtri, current_user)
    pdf  = render_pdf(tpl["definizione"], rows, ctx)
    nome = tpl["nome"].replace(" ", "_").lower()
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nome}.pdf"'})


@router.post("/anteprima", dependencies=[Depends(require_role(*RUOLI_DESIGNER))])
def anteprima_pdf(body: dict = Body(...), current_user: dict = Depends(get_current_user)):
    """Render PDF da una definizione non salvata (anteprima del designer)."""
    sorgente = body.get("sorgente")
    if sorgente not in CAMPI_SORGENTE:
        raise HTTPException(422, "Sorgente non valida")
    filtri = body.get("filtri") or {}
    rows = _carica_dati(sorgente, filtri, current_user)
    ctx  = _contesto(filtri, current_user)
    pdf  = render_pdf(body.get("definizione") or {}, rows, ctx)
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="anteprima.pdf"'})


@router.get("/excel/{sorgente}")
def genera_excel(
    sorgente: str,
    campagna_id:   Optional[int]  = Query(None),
    postazione_id: Optional[int]  = Query(None),
    data_da:       Optional[date] = Query(None),
    data_a:        Optional[date] = Query(None),
    stato:         Optional[str]  = Query(None),
    current_user:  dict           = Depends(get_current_user),
):
    filtri = _filtri_da_query(campagna_id, postazione_id, data_da, data_a, stato)
    rows = _carica_dati(sorgente, filtri, current_user)
    ctx  = _contesto(filtri, current_user)
    contenuto, nome = excel_per_sorgente(sorgente, rows, ctx)
    return StreamingResponse(io.BytesIO(contenuto),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'})
